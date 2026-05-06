import os
import tempfile
from io import BytesIO

import pandas as pd
import streamlit as st
from PIL import Image

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.cell import MergedCell


# =========================================================
# CONFIG
# =========================================================

st.set_page_config(
    page_title="WBMF-40AI Logistics Robot",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="collapsed"
)

TEMPLATE_FILE = "WBMF PO 1011953241 HAJU.xlsx"


# =========================================================
# ROBOTIC LOGISTICS THEME
# =========================================================

st.markdown(
    """
    <style>
    [data-testid="stSidebar"] {
        display: none;
    }

    [data-testid="collapsedControl"] {
        display: none;
    }

    .stApp {
        background:
            radial-gradient(circle at top left, rgba(0,255,170,0.13), transparent 28%),
            radial-gradient(circle at top right, rgba(0,162,255,0.13), transparent 30%),
            linear-gradient(135deg, #071018 0%, #0b1220 48%, #111827 100%);
        color: #e5f7ff;
    }

    .block-container {
        padding-top: 1.4rem;
        padding-bottom: 2rem;
        max-width: 1450px;
    }

    .robot-header {
        border: 1px solid rgba(0, 255, 170, 0.35);
        background: linear-gradient(135deg, rgba(13,27,42,0.94), rgba(17,24,39,0.94));
        border-radius: 22px;
        padding: 24px 28px;
        box-shadow: 0 0 28px rgba(0,255,170,0.12);
        margin-bottom: 20px;
    }

    .robot-title {
        font-size: 36px;
        font-weight: 900;
        letter-spacing: 0.5px;
        margin: 0;
        color: #dffcff;
    }

    .robot-subtitle {
        color: #92f7d5;
        font-size: 15px;
        margin-top: 8px;
    }

    .robot-badge {
        display: inline-block;
        padding: 7px 12px;
        border-radius: 999px;
        background: rgba(0,255,170,0.10);
        border: 1px solid rgba(0,255,170,0.35);
        color: #7cffd2;
        font-size: 13px;
        margin-right: 8px;
        margin-top: 14px;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 10px;
    }

    .stTabs [data-baseweb="tab"] {
        background: rgba(15,23,42,0.88);
        border: 1px solid rgba(56,189,248,0.25);
        border-radius: 14px;
        color: #c7f9ff;
        padding: 12px 18px;
    }

    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, rgba(0,255,170,0.20), rgba(56,189,248,0.20)) !important;
        border: 1px solid rgba(0,255,170,0.75) !important;
        color: #ffffff !important;
        box-shadow: 0 0 18px rgba(0,255,170,0.15);
    }

    [data-testid="stMetric"] {
        background: rgba(15,23,42,0.88);
        border: 1px solid rgba(0,255,170,0.25);
        border-radius: 16px;
        padding: 16px;
        box-shadow: 0 0 20px rgba(0,255,170,0.08);
    }

    .stButton > button {
        border-radius: 14px;
        border: 1px solid rgba(0,255,170,0.55);
        background: linear-gradient(135deg, #00b894, #0984e3);
        color: white;
        font-weight: 800;
        padding: 0.75rem 1rem;
        box-shadow: 0 0 22px rgba(0,255,170,0.18);
    }

    .stDownloadButton > button {
        border-radius: 14px;
        border: 1px solid rgba(0,255,170,0.55);
        background: linear-gradient(135deg, #10b981, #0ea5e9);
        color: white;
        font-weight: 800;
        padding: 0.75rem 1rem;
        box-shadow: 0 0 22px rgba(14,165,233,0.18);
    }

    [data-testid="stFileUploader"] {
        background: rgba(15,23,42,0.82);
        border: 1px dashed rgba(0,255,170,0.55);
        border-radius: 20px;
        padding: 20px;
        box-shadow: inset 0 0 18px rgba(0,255,170,0.06);
    }

    textarea {
        font-family: Consolas, monospace !important;
    }

    [data-testid="stAlert"] {
        border-radius: 15px;
    }

    h1, h2, h3 {
        color: #eaffff;
    }

    label, .stMarkdown, p, span {
        color: #d9f8ff;
    }

    .picture-panel {
        border: 1px solid rgba(0,255,170,0.35);
        background: rgba(15,23,42,0.78);
        border-radius: 22px;
        padding: 22px;
        margin-top: 10px;
        box-shadow: 0 0 26px rgba(0,255,170,0.09);
    }
    </style>
    """,
    unsafe_allow_html=True
)


# =========================================================
# HELPER FUNCTIONS
# =========================================================

def write_cell(ws, cell_address, value):
    ws[cell_address] = value if value not in [None, ""] else "-"


def set_cell_safe(ws, cell_address, value):
    """
    Tulis value ke cell tanpa merusak merged cell.
    """
    cell = ws[cell_address]

    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            top_left_cell = ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col
            )
            top_left_cell.value = value
            return


def clear_cell_safe(ws, cell_address):
    cell = ws[cell_address]

    if isinstance(cell, MergedCell):
        return

    cell.value = None


def clear_range_safe(ws, min_row, max_row, columns):
    for row in range(min_row, max_row + 1):
        for col in columns:
            clear_cell_safe(ws, f"{col}{row}")


def clear_picture_sheet(ws_picture, keep_title=True):
    """
    Bersihkan sheet PICTURE dari tulisan/gambar lama.
    Judul tetap dipertahankan.
    """
    ws_picture._images = []

    title_value = ws_picture["A1"].value if keep_title else None

    for row in ws_picture.iter_rows():
        for cell in row:
            if keep_title and cell.row == 1:
                continue

            if isinstance(cell, MergedCell):
                continue

            cell.value = None

    if keep_title:
        ws_picture["A1"] = title_value if title_value else "PICTURE & ATTACHMENT"


def find_row_by_text(ws, text):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == text:
                return cell.row
    return None


def safe_filename(text):
    text = str(text)
    invalid_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']

    for char in invalid_chars:
        text = text.replace(char, "-")

    return text


def normalize_qty(value, default=1):
    try:
        if value is None:
            return default

        value = str(value).strip()

        if value == "":
            return default

        value = value.replace(",", ".")

        number = float(value)

        if number.is_integer():
            number = int(number)

        return number

    except Exception:
        return default


def save_uploaded_image_to_temp(uploaded_file):
    """
    Simpan gambar upload tanpa mengubah resolusi asli.
    Yang diubah hanya ukuran tampilan object gambar di Excel.
    """
    uploaded_file.seek(0)

    file_ext = os.path.splitext(uploaded_file.name)[1].lower()

    if file_ext not in [".png", ".jpg", ".jpeg"]:
        file_ext = ".png"

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=file_ext)
    temp_file.write(uploaded_file.getbuffer())
    temp_file.close()

    return temp_file.name


def get_fitted_image_size(image_path, max_width=430, max_height=330):
    with Image.open(image_path) as img:
        original_width, original_height = img.size

    if original_width == 0 or original_height == 0:
        return max_width, max_height

    ratio = min(
        max_width / original_width,
        max_height / original_height
    )

    display_width = int(original_width * ratio)
    display_height = int(original_height * ratio)

    return display_width, display_height


def parse_material_direct_input(bulk_text):
    """
    Parse paste langsung dari Excel.

    Aturan:
    - Kolom pertama = Material Description.
    - Kolom kedua = Quantity jika ada dan angka.
    - Kolom ketiga dan seterusnya diabaikan.
    - Jika Quantity kosong/tidak valid, default = 1.
    """

    valid_items = []

    for line in bulk_text.splitlines():
        raw_line = line.strip()

        if raw_line == "":
            continue

        columns = [col.strip() for col in raw_line.split("\t")]

        if len(columns) == 1 and "," in raw_line:
            columns = [col.strip() for col in raw_line.split(",")]

        material_desc = columns[0].strip() if len(columns) >= 1 else ""

        if material_desc == "":
            continue

        quantity = 1

        if len(columns) >= 2:
            quantity = normalize_qty(columns[1], default=1)

        valid_items.append({
            "material_desc": material_desc,
            "quantity": quantity,
            "quantity_received": quantity,
            "job_site": "MACO MINING",
            "destination": "MACO HAULING",
            "uom": "EA"
        })

    return valid_items


def build_preview_dataframe(valid_items):
    rows = []

    for idx, item in enumerate(valid_items, start=1):
        rows.append({
            "No": idx,
            "Material Description": item["material_desc"],
            "Quantity": item["quantity"],
            "Quantity Received": item["quantity_received"],
            "Job Site": item["job_site"],
            "Destination": item["destination"],
            "UOM": item["uom"]
        })

    return pd.DataFrame(rows)


def write_header_direct_to_manifest_waybill(ws_manifest, ws_waybill, form_data):
    """
    Menulis header langsung ke Manifest dan Waybill agar Sheet1 bisa dihapus.
    Mapping ini menggantikan formula seperti =Sheet1!B1, =Sheet1!B2, dst.
    """

    # Manifest header
    set_cell_safe(ws_manifest, "E3", form_data["mf_number"])
    set_cell_safe(ws_manifest, "E4", form_data["po_sto"])
    set_cell_safe(ws_manifest, "E5", form_data["insurance_po_number"])
    set_cell_safe(ws_manifest, "E6", form_data["forwarder"])
    set_cell_safe(ws_manifest, "E7", form_data["delivery_mode"])
    set_cell_safe(ws_manifest, "E8", form_data["transportation_type"])
    set_cell_safe(ws_manifest, "E9", form_data["transportation_name"])
    set_cell_safe(ws_manifest, "E10", form_data["transportation_capacity"])
    set_cell_safe(ws_manifest, "E11", form_data["etd"])
    set_cell_safe(ws_manifest, "E12", form_data["eta"])
    set_cell_safe(ws_manifest, "E13", form_data["actual_arrival_date"])

    set_cell_safe(ws_manifest, "L3", form_data["from_location"])
    set_cell_safe(ws_manifest, "L8", form_data["to_location"])
    set_cell_safe(ws_manifest, "L11", form_data["attention"])
    set_cell_safe(ws_manifest, "L12", form_data["phone"])
    set_cell_safe(ws_manifest, "L13", form_data["company"])

    # Waybill header
    set_cell_safe(ws_waybill, "D3", form_data["wb_number"])


def replace_sheet1_formulas_with_values(wb, form_data):
    """
    Mengganti formula yang masih mengarah ke Sheet1 menjadi value langsung.
    Ini penting sebelum Sheet1 dihapus.
    """

    formula_value_map = {
        "=Sheet1!B1": form_data["mf_number"],
        "=Sheet1!B2": form_data["wb_number"],
        "=Sheet1!B3": form_data["po_sto"],
        "=Sheet1!B4": form_data["forwarder"],
        "=Sheet1!B5": form_data["delivery_mode"],
        "=Sheet1!B6": form_data["insurance_po_number"],
        "=Sheet1!B7": form_data["transportation_type"],
        "=Sheet1!B8": form_data["transportation_name"],
        "=Sheet1!B9": form_data["transportation_capacity"],
        "=Sheet1!B10": form_data["etd"],
        "=Sheet1!B11": form_data["eta"],
        "=Sheet1!B12": form_data["actual_arrival_date"],
        "=Sheet1!B13": form_data["from_location"],
        "=Sheet1!B14": form_data["to_location"],
        "=Sheet1!B15": form_data["attention"],
        "=Sheet1!B16": form_data["phone"],
        "=Sheet1!B17": form_data["company"],
    }

    for ws in wb.worksheets:
        if ws.title == "Sheet1":
            continue

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_normalized = (
                        cell.value
                        .replace("$", "")
                        .replace("'", "")
                        .strip()
                    )

                    if formula_normalized in formula_value_map:
                        cell.value = formula_value_map[formula_normalized]


def delete_sheet1_from_output(wb):
    """
    Hapus Sheet1 dari workbook output.
    """
    if "Sheet1" in wb.sheetnames:
        ws_sheet1 = wb["Sheet1"]
        wb.remove(ws_sheet1)

    if "Manifest" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Manifest")


# =========================================================
# GENERATE EXCEL
# =========================================================

def generate_excel(form_data, valid_items, uploaded_images):
    wb = load_workbook(TEMPLATE_FILE)

    ws_input = wb["Sheet1"]
    ws_manifest = wb["Manifest"]
    ws_waybill = wb["Waybill"]
    ws_picture = wb["PICTURE"]

    # =====================================================
    # UPDATE SHEET1 TEMPORARY
    # =====================================================
    # Sheet1 hanya dipakai sementara untuk menjaga kompatibilitas template.
    # Setelah header dipindah ke Manifest/Waybill, Sheet1 dihapus dari output.

    write_cell(ws_input, "B1", form_data["mf_number"])
    write_cell(ws_input, "B2", form_data["wb_number"])
    write_cell(ws_input, "B3", form_data["po_sto"])
    write_cell(ws_input, "B4", form_data["forwarder"])
    write_cell(ws_input, "B5", form_data["delivery_mode"])
    write_cell(ws_input, "B6", form_data["insurance_po_number"])
    write_cell(ws_input, "B7", form_data["transportation_type"])
    write_cell(ws_input, "B8", form_data["transportation_name"])
    write_cell(ws_input, "B9", form_data["transportation_capacity"])
    write_cell(ws_input, "B10", form_data["etd"])
    write_cell(ws_input, "B11", form_data["eta"])
    write_cell(ws_input, "B12", form_data["actual_arrival_date"])
    write_cell(ws_input, "B13", form_data["from_location"])
    write_cell(ws_input, "B14", form_data["to_location"])
    write_cell(ws_input, "B15", form_data["attention"])
    write_cell(ws_input, "B16", form_data["phone"])
    write_cell(ws_input, "B17", form_data["company"])

    # =====================================================
    # HEADER LANGSUNG KE MANIFEST / WAYBILL
    # =====================================================

    write_header_direct_to_manifest_waybill(
        ws_manifest=ws_manifest,
        ws_waybill=ws_waybill,
        form_data=form_data
    )

    # =====================================================
    # UPDATE MANIFEST
    # =====================================================

    manifest_start_row = 17
    total_qty_row = find_row_by_text(ws_manifest, "Total Quantity")
    manifest_end_row = total_qty_row - 1 if total_qty_row else 130

    clear_range_safe(
        ws_manifest,
        manifest_start_row,
        manifest_end_row,
        ["A", "B", "C", "I", "K", "P"]
    )

    manifest_row = manifest_start_row
    total_quantity = 0
    manifest_inserted = 0

    for idx, item in enumerate(valid_items, start=1):
        if manifest_row > manifest_end_row:
            break

        set_cell_safe(ws_manifest, f"A{manifest_row}", idx)
        set_cell_safe(ws_manifest, f"B{manifest_row}", form_data["wb_number"])
        set_cell_safe(ws_manifest, f"C{manifest_row}", item["material_desc"])
        set_cell_safe(ws_manifest, f"I{manifest_row}", item["quantity"])
        set_cell_safe(ws_manifest, f"K{manifest_row}", item["destination"])
        set_cell_safe(ws_manifest, f"P{manifest_row}", item["uom"])

        total_quantity += item["quantity"]
        manifest_inserted += 1

        # Template Manifest memakai jarak 2 row antar item.
        manifest_row += 2

    if total_qty_row:
        set_cell_safe(ws_manifest, f"I{total_qty_row}", total_quantity)

    # =====================================================
    # UPDATE WAYBILL
    # =====================================================

    waybill_start_row = 8
    prepared_row = find_row_by_text(ws_waybill, "Prepared By")
    waybill_end_row = prepared_row - 1 if prepared_row else 30

    clear_range_safe(
        ws_waybill,
        waybill_start_row,
        waybill_end_row,
        ["A", "C", "H", "K", "M", "O"]
    )

    waybill_row = waybill_start_row
    waybill_inserted = 0

    for idx, item in enumerate(valid_items, start=1):
        if waybill_row > waybill_end_row:
            break

        set_cell_safe(ws_waybill, f"A{waybill_row}", idx)
        set_cell_safe(ws_waybill, f"C{waybill_row}", item["material_desc"])
        set_cell_safe(ws_waybill, f"H{waybill_row}", item["job_site"])
        set_cell_safe(ws_waybill, f"K{waybill_row}", item["destination"])
        set_cell_safe(ws_waybill, f"M{waybill_row}", item["quantity"])
        set_cell_safe(ws_waybill, f"O{waybill_row}", item["quantity_received"])

        waybill_inserted += 1
        waybill_row += 1

    # =====================================================
    # UPDATE PICTURE
    # =====================================================

    clear_picture_sheet(ws_picture, keep_title=True)

    if uploaded_images:
        picture_slots = [
            {"cell": "B3", "max_width": 430, "max_height": 330},
            {"cell": "K3", "max_width": 430, "max_height": 330},
            {"cell": "B31", "max_width": 430, "max_height": 330},
            {"cell": "K31", "max_width": 430, "max_height": 330},
            {"cell": "B59", "max_width": 430, "max_height": 330},
            {"cell": "K59", "max_width": 430, "max_height": 330},
            {"cell": "B87", "max_width": 430, "max_height": 330},
            {"cell": "K87", "max_width": 430, "max_height": 330},
        ]

        for uploaded_file, slot in zip(uploaded_images, picture_slots):
            temp_image_path = save_uploaded_image_to_temp(uploaded_file)

            img = XLImage(temp_image_path)

            display_width, display_height = get_fitted_image_size(
                temp_image_path,
                max_width=slot["max_width"],
                max_height=slot["max_height"]
            )

            img.width = display_width
            img.height = display_height

            ws_picture.add_image(img, slot["cell"])

    # =====================================================
    # HAPUS SHEET1 DARI OUTPUT
    # =====================================================
    # Sebelum Sheet1 dihapus, semua formula yang masih refer ke Sheet1 diganti value.

    replace_sheet1_formulas_with_values(wb, form_data)
    delete_sheet1_from_output(wb)

    # Recalculate tidak terlalu dibutuhkan lagi karena formula Sheet1 sudah diganti value,
    # tapi tetap aman untuk formula lain di workbook.
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return {
        "file": output,
        "total_items_input": len(valid_items),
        "total_quantity": total_quantity,
        "manifest_inserted": manifest_inserted,
        "waybill_inserted": waybill_inserted
    }


# =========================================================
# STREAMLIT UI
# =========================================================

st.markdown(
    """
    <div class="robot-header">
        <p class="robot-title">🤖 WBMF-40AI Logistics Robot</p>
        <div class="robot-subtitle">
            Direct Excel Paste to Manifest, Waybill & Picture Attachment Generator
        </div>
        <span class="robot-badge">📦 Direct Manifest Input</span>
        <span class="robot-badge">🚚 Direct Waybill Input</span>
        <span class="robot-badge">🧠 Sheet1 Removed on Output</span>
        <span class="robot-badge">🖼️ Auto Picture Clean</span>
    </div>
    """,
    unsafe_allow_html=True
)

if not os.path.exists(TEMPLATE_FILE):
    st.error(f"Template Excel tidak ditemukan: {TEMPLATE_FILE}")
    st.info("Pastikan file template Excel ada satu folder dengan app.py.")
    st.stop()


tab1, tab2, tab3 = st.tabs([
    "🤖 Header Manifest / Waybill",
    "📦 Direct Material Paste",
    "🖼️ Picture Bay"
])


# =========================================================
# TAB 1 - HEADER
# =========================================================

with tab1:
    st.subheader("🤖 Header Manifest / Waybill")
    st.caption("Input ini langsung ditulis ke Manifest dan Waybill. Sheet1 akan dihapus dari file download.")

    col1, col2 = st.columns(2)

    with col1:
        mf_number = st.text_input(
            "MF Number",
            value="MF40AI-02052026/MACOMINING/02"
        )

        wb_number = st.text_input(
            "WB Number",
            value="WB40AI-02052026/MACOMINING/02"
        )

        po_sto = st.text_input(
            "PO / STO",
            value="1011953241"
        )

        forwarder = st.text_input(
            "Forwarder / Pengirim",
            value="-"
        )

        delivery_mode = st.selectbox(
            "Delivery Mode",
            ["LAND", "SEA", "AIR"],
            index=0
        )

        insurance_po_number = st.text_input(
            "Insurance PO Number",
            value="-"
        )

        transportation_type = st.text_input(
            "Transportation Type",
            value="-"
        )

        transportation_name = st.text_input(
            "Transportation Name",
            value="-"
        )

        transportation_capacity = st.text_input(
            "Transportation Capacity",
            value="-"
        )

    with col2:
        etd = st.text_input(
            "Estimated Time of Departure",
            value="May, 02 2026"
        )

        eta = st.text_input(
            "Estimated Time of Arrival",
            value="May, 02 2026"
        )

        actual_arrival_date = st.text_input(
            "Actual Arrival Date",
            value="-"
        )

        from_location = st.text_area(
            "From",
            value="PT. SAPTAINDRA SEJATI SITE MACO MINING"
        )

        to_location = st.text_area(
            "To",
            value="PT. SAPTAINDRA SEJATI SITE MACO HAULING"
        )

        attention = st.text_input(
            "Attention / Penerima",
            value="Bapak Fachry"
        )

        phone = st.text_input(
            "Phone",
            value="+62 812-8347-1699"
        )

        company = st.text_input(
            "Company",
            value="PT SAPTAINDRA SEJATI"
        )


# =========================================================
# TAB 2 - DIRECT MATERIAL PASTE
# =========================================================

with tab2:
    st.subheader("📦 Direct Material Paste")
    st.caption(
        "Paste langsung dari Excel. Sistem mengambil kolom pertama sebagai Material dan kolom kedua sebagai Quantity jika ada. "
        "Kolom lain otomatis diabaikan."
    )

    bulk_material_text = st.text_area(
        "Paste Material dari Excel di sini",
        height=330,
        placeholder=(
            "Contoh 1 kolom:\\n"
            "24122002-0008\\n"
            "4T-32032X\\n"
            "4T-32036XE1PX4\\n"
            "561-22-62970\\n\\n"
            "Contoh 2 kolom:\\n"
            "24122002-0008    5\\n"
            "4T-32032X        2\\n"
            "4T-32036XE1PX4   1"
        )
    )

    valid_items_preview = parse_material_direct_input(bulk_material_text)
    preview_df = build_preview_dataframe(valid_items_preview)

    total_preview_qty = sum(item["quantity"] for item in valid_items_preview)

    col_a, col_b, col_c = st.columns(3)

    with col_a:
        st.metric("Total Item Detected", len(valid_items_preview))

    with col_b:
        st.metric("Total Quantity", total_preview_qty)

    with col_c:
        st.metric("Output Sheets", "3 Sheets")

    if not preview_df.empty:
        st.markdown("### Preview yang akan masuk ke Manifest / Waybill")
        st.dataframe(
            preview_df,
            use_container_width=True,
            hide_index=True
        )
    else:
        st.info("Belum ada material yang terdeteksi. Paste data material dari Excel terlebih dahulu.")


# =========================================================
# TAB 3 - PICTURE
# =========================================================

with tab3:
    st.subheader("🖼️ Picture Bay - Logistics Evidence Upload")
    st.caption("Sheet PICTURE otomatis dibersihkan. Upload gambar akan masuk ke slot attachment.")

    st.markdown('<div class="picture-panel">', unsafe_allow_html=True)

    uploaded_images = st.file_uploader(
        "Drop / Browse picture attachment",
        type=["png", "jpg", "jpeg"],
        accept_multiple_files=True,
        help="Upload foto material, packaging, evidence delivery, atau attachment lain."
    )

    if uploaded_images:
        st.success(f"{len(uploaded_images)} gambar siap diproses oleh Logistics Robot.")

        preview_cols = st.columns(4)

        for idx, uploaded_image in enumerate(uploaded_images):
            with preview_cols[idx % 4]:
                st.image(
                    uploaded_image,
                    caption=uploaded_image.name,
                    use_container_width=True
                )
    else:
        st.info("Picture optional. Jika tidak ada gambar, sheet PICTURE tetap dibersihkan dari data/gambar lama.")

    st.markdown('</div>', unsafe_allow_html=True)


# =========================================================
# GENERATE
# =========================================================

st.divider()

form_data = {
    "mf_number": mf_number,
    "wb_number": wb_number,
    "po_sto": po_sto,
    "forwarder": forwarder,
    "delivery_mode": delivery_mode,
    "insurance_po_number": insurance_po_number,
    "transportation_type": transportation_type,
    "transportation_name": transportation_name,
    "transportation_capacity": transportation_capacity,
    "etd": etd,
    "eta": eta,
    "actual_arrival_date": actual_arrival_date,
    "from_location": from_location,
    "to_location": to_location,
    "attention": attention,
    "phone": phone,
    "company": company
}

generate = st.button(
    "🤖 Generate Excel - Sheet1 Removed",
    type="primary",
    use_container_width=True
)

if generate:
    valid_items = parse_material_direct_input(bulk_material_text)

    if len(valid_items) == 0:
        st.warning("Minimal paste 1 material terlebih dahulu.")
    else:
        try:
            result = generate_excel(
                form_data=form_data,
                valid_items=valid_items,
                uploaded_images=uploaded_images
            )

            safe_po_sto = safe_filename(po_sto)
            safe_wb_number = safe_filename(wb_number)

            excel_filename = f"WBMF_{safe_po_sto}_{safe_wb_number}.xlsx"

            st.success(
                f"Mission complete. Excel berhasil dibuat tanpa Sheet1. "
                f"Input item: {result['total_items_input']}, "
                f"Manifest inserted: {result['manifest_inserted']}, "
                f"Waybill inserted: {result['waybill_inserted']}, "
                f"Total quantity: {result['total_quantity']}"
            )

            if result["manifest_inserted"] < result["total_items_input"]:
                st.warning(
                    "Sebagian item tidak masuk ke Manifest karena area row pada template Manifest sudah penuh."
                )

            if result["waybill_inserted"] < result["total_items_input"]:
                st.warning(
                    "Sebagian item tidak masuk ke Waybill karena area row pada template Waybill sudah penuh."
                )

            st.download_button(
                label="⬇️ Download Excel Output",
                data=result["file"],
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

        except Exception as error:
            st.error("Generate file gagal.")
            st.code(str(error))